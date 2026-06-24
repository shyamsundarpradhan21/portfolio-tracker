#!/usr/bin/env python3
"""
parse-broker-tax.py — reusable realized-P&L extractor.

Reads every broker tax/P&L report dropped in data/reports/ (gitignored, transient)
and emits data/broker-tax.json — a committed, PII-stripped canonical store so the
app's realized figures are *derived from the reports*, never hand-curated/stale,
and survive without re-uploading the raw files.

WHAT IT NEVER WRITES OUT: PAN, client IDs, account numbers, names, e-mail, or the
raw trade-by-trade list. Only per-FY / per-segment realized totals, trade counts,
and the top movers (symbol + amount) that the UI already displays.

Workflow:  drop reports in data/reports/  ->  python scripts/parse-broker-tax.py
           ->  git add data/broker-tax.json  ->  node scripts/seed-portfolio-kv.mjs

Supports: Zerodha Console tax-P&L (.xlsx, one file per FY), Fyers tax-P&L (.csv,
one per FY), Dhan tax report (.xls, all-time multi-segment), Vested/DriveWealth
P&L statement (.xlsx). Upstox (.xlsx) slots in once uploaded (TODO marker below).

Indian FY = 1 Apr – 31 Mar.  Label form: FY25-26.
"""
import os, re, csv, glob, json, datetime as dt
import openpyxl

REPORTS = os.path.join("data", "reports")
OUT = os.path.join("data", "broker-tax.json")

# ── account → role map (output uses these labels, never the client IDs) ──────
ACCOUNTS = {
    "GWS919":  {"broker": "zerodha", "owner": "mom",  "sleeve": "indian_equity", "label": "zerodha_mom"},
    "YXA918":  {"broker": "zerodha", "owner": "self", "sleeve": "trading",       "label": "zerodha_self"},
    "YS59535": {"broker": "fyers",   "owner": "self", "sleeve": "trading",       "label": "fyers"},
    # Dhan keyed by UCC, Vested by DriveWealth acc — matched on file content/type below.
}

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

def fy_of(d):
    """datetime/date -> 'FY25-26'."""
    if d is None:
        return None
    y, m = d.year, d.month
    s = y if m >= 4 else y - 1
    return f"FY{str(s)[-2:]}-{str(s + 1)[-2:]}"

def fy_from_pair(a, b):
    """2025,2026 -> 'FY25-26'."""
    return f"FY{str(a)[-2:]}-{str(b)[-2:]}"

def num(x):
    """Best-effort float; '' / None / non-numeric -> None."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", "")
    if s in ("", "-", "NA", "N/A"):
        return None
    try:
        return float(s)
    except ValueError:
        return None

def parse_date(x):
    if x is None or x == "":
        return None
    if isinstance(x, (dt.datetime, dt.date)):
        return x
    s = str(x).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y", "%d %b  %Y"):
        try:
            return dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    # Zerodha tradewise uses ISO; Fyers '01 Apr  2025' (double space)
    s2 = re.sub(r"\s+", " ", s)
    try:
        return dt.datetime.strptime(s2, "%d %b %Y")
    except ValueError:
        return None

# ── Column detection BY HEADER NAME — robust to per-broker / per-segment layouts
# (F&O rows carry expiry/strike, shifting the date & P&L columns vs equity). ──
def _find_col(hdr, *subs):
    for i, c in enumerate(hdr):
        cl = str(c or "").strip().lower()
        if cl and any(s in cl for s in subs):
            return i
    return None

def _date_col(hdr):
    c = _find_col(hdr, "sell date", "exit date", "sold date", "close date")
    if c is None:                              # fall back to any 'date' that isn't the buy/open one
        c = next((j for j, x in enumerate(hdr)
                  if "date" in str(x or "").lower() and not any(
                      b in str(x or "").lower() for b in ("buy", "open", "acqu", "entry"))), None)
    return c

def _pnl_col(hdr):
    return _find_col(hdr, "realized p&l", "realised p&l", "net p&l", "net pnl",
                     "realized", "realised", "pnl", "p&l", "profit")

def _peekrow(r):
    out = []
    for i, c in enumerate(r):
        if c in (None, ""):
            continue
        s = str(c).strip()
        if num(s) is not None:
            s = "#"
        elif re.match(r"^[A-Z]{5}\d{4}[A-Z]$", s):
            s = "<id>"
        elif len(s) > 20:
            s = s[:18] + "…"
        out.append(f"{i}:{s}")
    return out

# Bucket a 'Tradewise Exits'-style sheet's F&O sections into date -> realised P&L,
# detecting the date & P&L columns by header name. Returns (by_day, peek_rows).
def tradewise_fno_daily(tw):
    out, peek = {}, []
    in_fno, date_c, pnl_c = False, None, None
    for r in tw:
        nonempty = [c for c in r if c not in (None, "")]
        first = str(nonempty[0]).strip() if nonempty else ""
        fl = first.lower()
        if len(nonempty) <= 2 and first:               # a section title / metadata row
            if ("f&o" in fl or "future" in fl or "option" in fl) and "curren" not in fl and "commod" not in fl:
                in_fno, date_c, pnl_c = True, None, None
            elif "symbol" not in fl:
                in_fno = False
            continue
        if not in_fno:
            continue
        if date_c is None:                             # the F&O sub-table header row
            dc, pc = _date_col(r), _pnl_col(r)
            if dc is not None and pc is not None:
                date_c, pnl_c = dc, pc
            elif len(peek) < 5:
                peek.append(_peekrow(r))
            continue
        d = parse_date(r[date_c]) if date_c < len(r) else None
        p = num(r[pnl_c]) if pnl_c < len(r) else None
        if d and p is not None:
            out[d.date().isoformat()] = round(out.get(d.date().isoformat(), 0.0) + p, 2)
        elif len(peek) < 5:
            peek.append(_peekrow(r))
    return out, peek

def top_movers(by_sym, k=3, usd=False):
    rnd = (lambda a: round(a, 2)) if usd else (lambda a: round(a))
    items = sorted(by_sym.items(), key=lambda kv: kv[1], reverse=True)
    win = [{"sym": s, "amt": rnd(a)} for s, a in items if a > 0][:k]
    los = [{"sym": s, "amt": rnd(a)} for s, a in items[::-1] if a < 0][:k]
    return win, los

def fmt_asof(d):
    return f"{d.day:02d} {MONTHS[d.month - 1]} {d.year}" if d else None

# ── label-based cell lookup for the Zerodha / Fyers summary blocks ───────────
def find_value_after(rows, label_substr):
    """Find first cell containing label_substr; return the next non-empty cell to its right."""
    low = label_substr.lower()
    for r in rows:
        for i, c in enumerate(r):
            if c is not None and low in str(c).strip().lower():
                for j in range(i + 1, len(r)):
                    v = num(r[j])
                    if v is not None:
                        return v
    return None

# ── Zerodha Console tax-P&L (.xlsx, one file per FY) ─────────────────────────
def parse_zerodha(path):
    m = re.search(r"taxpnl-([A-Z0-9]+)-(\d{4})_(\d{4})", os.path.basename(path))
    if not m:
        return None
    acct_id, y1, y2 = m.group(1), int(m.group(2)), int(m.group(3))
    meta = ACCOUNTS.get(acct_id, {"label": f"zerodha_{acct_id}", "sleeve": "trading", "owner": "self"})
    fy = fy_from_pair(y1, y2)
    wb = openpyxl.load_workbook(path, data_only=True)

    def sheet_rows(prefix):
        names = [s for s in wb.sheetnames if s.startswith(prefix)]
        if not names:
            return []
        return [[c.value for c in row] for row in wb[names[0]].iter_rows()]

    eq = sheet_rows("Equity and Non Equity")
    stcg = find_value_after(eq, "Short Term profit") or 0.0
    ltcg = find_value_after(eq, "Long Term profit") or 0.0
    intra = find_value_after(eq, "Intraday/Speculative profit") or 0.0
    noneq = find_value_after(eq, "Non Equity profit") or 0.0   # ETFs (e.g. GOLDBEES), SGBs
    fno_rows = sheet_rows("F&O")
    opt = find_value_after(fno_rows, "Options Realized Profit") or 0.0
    fut = find_value_after(fno_rows, "Futures Realized Profit") or 0.0

    # Tradewise Exits: per-symbol realized, EQUITY sections only. The sheet also
    # carries 'F&O' / 'Non Equity' / 'Currency' / 'Commodity' / 'Mutual Funds'
    # sections (e.g. NIFTY options for a self account) — those must NOT leak into
    # the equity winners/losers. Track the section; aggregate only under 'Equity*'.
    # n = distinct (symbol, exit-date) = positions closed (not lot rows).
    tw = sheet_rows("Tradewise Exits")
    by_sym, exits, last_exit = {}, set(), None
    section, in_table = None, False
    for r in tw:                      # EQUITY (+ ETF/SGB) per-symbol — F&O handled below
        c1 = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        if not c1:
            continue
        if c1 == "Symbol":            # column header → trade rows follow
            in_table = True
            continue
        profit = num(r[8]) if len(r) > 8 else None
        if profit is None:            # a section title (or metadata) → new section
            section, in_table = c1, False
            continue
        if not in_table or not section or not (section.startswith("Equity") or section == "Non Equity"):
            continue                  # equity + ETF/SGB only; skip F&O / currency / commodity / MF
        by_sym[c1] = by_sym.get(c1, 0.0) + profit
        ed = parse_date(r[4]) if len(r) > 4 else None
        exits.add((c1, ed.date() if ed else None))
        if ed and (last_exit is None or ed > last_exit):
            last_exit = ed

    # F&O daily — header-name detection over the Tradewise F&O sections (their
    # columns differ from equity: expiry/strike shift the date & P&L positions).
    fno_by_day, fno_peek = tradewise_fno_daily(tw)
    for k in fno_by_day:
        ed = parse_date(k)
        if ed and (last_exit is None or ed > last_exit):
            last_exit = ed

    # Sanity: F&O daily should reconcile to the F&O sheet's options+futures total.
    ssum = round(opt + fut)
    if fno_by_day:
        dsum = round(sum(fno_by_day.values()))
        if abs(dsum - ssum) > max(50, abs(ssum) * 0.02):
            print(f"  ! {meta['label']} {fy}: F&O daily Σ{dsum} != summary {ssum}")
    elif ssum != 0:
        print(f"  ! {meta['label']} {fy}: F&O summary ₹{ssum} but no daily rows matched.")
        for pr in fno_peek:
            print(f"      {'  '.join(pr)}")

    return {
        "label": meta["label"], "owner": meta["owner"], "sleeve": meta["sleeve"], "fy": fy,
        "equity": {"stcg": round(stcg), "ltcg": round(ltcg), "intraday": round(intra),
                   "nonequity": round(noneq), "realized": round(stcg + ltcg + intra + noneq)},
        "fno": {"options": round(opt), "futures": round(fut), "realized": round(opt + fut)},
        "fnoByDay": fno_by_day,
        "n": len(exits), "by_sym": by_sym, "lastExit": last_exit,
    }

# ── Fyers tax-P&L (.csv, one file per FY) ────────────────────────────────────
def parse_fyers(path):
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        rows = list(csv.reader(f))
    acct = None
    fy = None
    for r in rows:
        if len(r) >= 2 and str(r[0]).strip() == "Client ID":
            acct = str(r[1]).strip()
        if len(r) >= 2 and str(r[0]).strip() == "Financial Year" and "-" in str(r[1]):
            a, b = str(r[1]).split("-")
            fy = fy_from_pair(int(a), int(b))
        if acct and fy:
            break
    meta = ACCOUNTS.get(acct, {"label": "fyers", "sleeve": "trading", "owner": "self"})
    realized = find_value_after(rows, "Realised P&L Summary") or 0.0
    opt = find_value_after(rows, "Taxable Options P&L") or 0.0
    fut = find_value_after(rows, "Taxable Future P&L") or 0.0
    stcg = find_value_after(rows, "Net STCG P&L") or 0.0
    ltcg = find_value_after(rows, "Net LTCG P&L") or 0.0
    # Per-trade F&O daily: the report lists each closed trade with col1 = P&L,
    # col3 = Segment ('Derivatives' for F&O, 'Equity' otherwise), col8 = Sell date.
    # Bucket the Derivatives rows by their exact sell-date for the calendar backfill.
    # (col1 is realised P&L per trade; treated as gross — Fyers gives no per-trade
    # charge split. Equity rows are skipped so only F&O lands in the F&O calendar.)
    last = None
    fno_by_day = {}
    for r in rows:
        if len(r) <= 8:
            continue
        d = parse_date(r[8])
        if d and (last is None or d > last):
            last = d
        seg = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""
        pnl = num(r[1])
        if d and pnl is not None and seg == "Derivatives":
            k = d.date().isoformat()
            fno_by_day[k] = round(fno_by_day.get(k, 0.0) + pnl, 2)
    # Sanity: the daily sum should reconcile to the FY F&O summary (options+futures).
    if fno_by_day:
        dsum, ssum = round(sum(fno_by_day.values())), round(opt + fut)
        if abs(dsum - ssum) > max(50, abs(ssum) * 0.02):
            print(f"  ! fyers {fy}: F&O daily Σ{dsum} != summary {ssum} (check Segment/P&L cols)")
    return {
        "label": meta["label"], "owner": meta["owner"], "sleeve": meta["sleeve"], "fy": fy,
        "equity": {"stcg": round(stcg), "ltcg": round(ltcg), "intraday": 0,
                   "realized": round(stcg + ltcg)},
        "fno": {"options": round(opt), "futures": round(fut), "realized": round(opt + fut)},
        "fnoByDay": fno_by_day,
        "realizedNet": round(realized), "lastExit": last,
    }

# ── Dhan tax report (.xls, all-time) — segment-summary Gross P&L (FIFO) ───────
# Per-FY bucketing is unreliable from this export; the segment summary rows
# (Intraday/Short Term/Long Term ; F&O Summary) are authoritative all-time totals.
# Per-FY trading P&L lives in fno-ledger.json from the live sync — this is a check.
def parse_dhan(path):
    import xlrd
    wb = xlrd.open_workbook(path)
    out = {}
    fno_by_fy = {}
    fno_by_day = {}
    for seg, sheet in (("equity", "Equity"), ("fno", "F&O")):
        if sheet not in wb.sheet_names():
            continue
        ws = wb.sheet_by_name(sheet)
        rows = [[ws.cell_value(ri, ci) for ci in range(ws.ncols)] for ri in range(ws.nrows)]
        # all-time segment summary (Gross P&L of the summary rows above the per-trade table)
        gp_col = None
        total = 0.0
        for r in rows:
            if gp_col is None:
                for i, c in enumerate(r):
                    if c is not None and str(c).strip() == "Gross P&L":
                        gp_col = i
                continue
            c0 = str(r[0]).strip() if r[0] is not None else ""
            if num(c0) is not None:           # reached per-trade Sr rows
                break
            gp = num(r[gp_col]) if gp_col < len(r) else None
            if c0 and gp is not None:          # a segment summary row
                total += gp
            elif c0 and gp is None:            # section title (e.g. 'Tradewise Details')
                break
        out[seg] = round(total)
        # F&O per-FY AND per-day: bucket each tradewise row by Sell-Date. The day
        # bucket (date → gross+net) is what backfills the Trading-tab daily calendar
        # via scripts/backfill-fno-ledger.mjs; it is non-PII (date + ₹ only, no
        # symbol/account), same privacy class as the per-FY totals.
        if seg == "fno":
            sd = gpc = npc = None
            for r in rows:
                for i, c in enumerate(r):
                    s = str(c).strip() if c is not None else ""
                    if s == "Sell Date": sd = i
                    elif s == "Gross P&L": gpc = i
                    elif s == "Net P&L": npc = i
                if sd is not None and gpc is not None and npc is not None:
                    break
            for r in rows:
                if num(r[0]) is None:          # only Sr-numbered trade rows
                    continue
                sell = parse_date(r[sd]) if sd is not None and sd < len(r) else None
                g = num(r[gpc]) if gpc is not None and gpc < len(r) else None
                nv = num(r[npc]) if npc is not None and npc < len(r) else None
                if sell is None or g is None:
                    continue
                b = fno_by_fy.setdefault(fy_of(sell), {"gross": 0.0, "net": 0.0})
                b["gross"] += g
                b["net"] += nv if nv is not None else g
                day = fno_by_day.setdefault(sell.date().isoformat(), {"gross": 0.0, "net": 0.0})
                day["gross"] += g
                day["net"] += nv if nv is not None else g
    return {"label": "dhan", "owner": "self", "sleeve": "trading", "allTime": out,
            "fnoByFY": {k: {"gross": round(v["gross"]), "net": round(v["net"])}
                        for k, v in fno_by_fy.items()},
            "fnoByDay": {k: {"gross": round(v["gross"], 2), "net": round(v["net"], 2)}
                         for k, v in sorted(fno_by_day.items())}}

# ── Vested/DriveWealth P&L statement (.xlsx) — Realized breakdown by Date Sold ─
def parse_vested(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sn = [s for s in wb.sheetnames if s.strip().startswith("Realized P&L - Breakdown")]
    if not sn:
        return None
    rows = [[c.value for c in r] for r in wb[sn[0]].iter_rows()]
    # header: Security, Quantity, Date Sold, ..., Profit/Loss (USD)[8], ..., P/L(INR)[11]
    by_fy = {}        # fy -> {'usd':x,'inr':y,'exits':set,'sym':{...}}
    last = None
    for r in rows[1:]:
        sec = str(r[0]).strip() if r[0] else ""
        sold = parse_date(r[2]) if len(r) > 2 else None
        usd = num(r[8]) if len(r) > 8 else None
        inr = num(r[11]) if len(r) > 11 else None
        if not sec or sold is None or usd is None:
            continue
        fy = fy_of(sold)
        b = by_fy.setdefault(fy, {"usd": 0.0, "inr": 0.0, "exits": set(), "sym": {}})
        b["usd"] += usd
        b["inr"] += inr or 0.0
        b["exits"].add((sec, sold.date()))     # distinct sell events, not lots
        b["sym"][sec] = b["sym"].get(sec, 0.0) + usd
        if last is None or sold > last:
            last = sold
    return {"by_fy": by_fy, "lastExit": last}

# ── Upstox realized P&L (zipped .xlsx, one per FY) — F&O trading (UCC 7BB93B) ──
def parse_upstox_zip(path):
    import zipfile, io
    fy = None
    m = re.search(r"realizedPnL_(\d{4})_", os.path.basename(path))
    if m:
        s = m.group(1)
        fy = f"FY{s[:2]}-{s[2:]}"
    with zipfile.ZipFile(path) as zf:
        xlsx = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not xlsx:
            return None        # some FYs are pdf-only — the xlsx twin carries the data
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(xlsx[0])), data_only=True)
    rows = [[c.value for c in r] for r in wb[wb.sheetnames[0]].iter_rows()]
    net = find_value_after(rows, "Net P&L")
    if net is None or round(net) == 0:
        return None        # empty FY (no F&O activity that year) — skip the noise
    gross = find_value_after(rows, "Gross P&L")

    # Per-trade daily F&O: the REALIZED_PNL sheet has a trade table. Detect its
    # date & P&L columns BY HEADER NAME (layout-agnostic), then bucket by sell date.
    # The whole account is F&O (UCC 7BB93B), so no equity filtering is needed.
    fno_by_day = {}
    hdr_i = date_c = pnl_c = None
    for i, r in enumerate(rows):
        dc, pc = _date_col(r), _pnl_col(r)
        if dc is not None and pc is not None:
            hdr_i, date_c, pnl_c = i, dc, pc
            break
    if hdr_i is not None:
        for r in rows[hdr_i + 1:]:
            d = parse_date(r[date_c]) if date_c < len(r) else None
            p = num(r[pnl_c]) if pnl_c < len(r) else None
            if d and p is not None:
                k = d.date().isoformat()
                fno_by_day[k] = round(fno_by_day.get(k, 0.0) + p, 2)
    if fno_by_day:
        dsum = round(sum(fno_by_day.values()))
        if abs(dsum - round(net)) > max(50, abs(round(net)) * 0.02):
            print(f"  ! upstox {fy}: daily Σ{dsum} != Net P&L {round(net)} (date c{date_c}, pnl c{pnl_c})")
    elif round(net) != 0:
        print(f"  ! upstox {fy}: Net P&L ₹{round(net)} but no per-trade table found. First rows:")
        for r in rows[:8]:
            pr = _peekrow(r)
            if pr:
                print(f"      {'  '.join(pr)}")
    return {"label": "upstox", "owner": "self", "sleeve": "trading", "fy": fy,
            "fno": {"realized": round(net), "gross": round(gross) if gross is not None else None},
            "fnoByDay": fno_by_day}

# ── Astha Credit & Securities F&O P&L (.csv, one per FY) — old F&O trading ─────
def parse_astha(path):
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        rows = list(csv.reader(f))
    fy = None
    for r in rows[:8]:
        if len(r) >= 2 and str(r[0]).strip() == "Report Date":
            m = re.search(r"(\d{4})-\d\d-\d\d to (\d{4})", str(r[1]))
            if m:
                fy = fy_from_pair(int(m.group(1)), int(m.group(2)))
    tot, pcol = 0.0, None
    for r in rows:
        if pcol is None:
            for i, c in enumerate(r):
                if c is not None and str(c).strip() == "P&L":
                    pcol = i
            continue
        v = num(r[pcol]) if pcol < len(r) else None
        if v is not None:
            tot += v
    return {"label": "astha", "owner": "self", "sleeve": "trading", "fy": fy,
            "fno": {"realized": round(tot)}}

# ── Groww equity P&L (.xlsx) — self equity, per-FY via Trade Level (0258131546) ─
def parse_groww(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    tl = [[c.value for c in r] for r in wb["Trade Level"].iter_rows()] if "Trade Level" in wb.sheetnames else []
    realized = find_value_after(tl, "Realised P&L")    # 'Realised P&L' row precedes 'Unrealised'
    # 'Realised trades' table: Stock[0] ISIN[1] Qty[2] BuyDate[3] BuyPrice[4] BuyValue[5]
    #                          SellDate[6] SellPrice[7] SellValue[8]  → P&L = SellValue − BuyValue
    by_fy = {}        # fy -> {amt, exits:set, by_sym:{}}
    hdr = False
    for r in tl:
        c0 = str(r[0]).strip() if r and r[0] is not None else ""
        if c0 == "Stock name":
            hdr = True
            continue
        if not hdr or not c0 or c0.lower() == "total":
            continue
        sell = parse_date(r[6]) if len(r) > 6 else None
        bv = num(r[5]) if len(r) > 5 else None
        sv = num(r[8]) if len(r) > 8 else None
        if sell is None or bv is None or sv is None:
            continue
        pnl = sv - bv
        d = by_fy.setdefault(fy_of(sell), {"amt": 0.0, "exits": set(), "by_sym": {}})
        d["amt"] += pnl
        d["exits"].add((c0, sell.date()))
        d["by_sym"][c0] = d["by_sym"].get(c0, 0.0) + pnl
    return {"label": "groww", "owner": "self", "sleeve": "indian_equity_self",
            "allTime": round(realized) if realized is not None else None, "by_fy": by_fy}

# ── build the RealizedPanel-shaped block from per-FY accumulators ─────────────
def realized_block(per_fy, asof, source, usd=False):
    """per_fy: {fy: {'amt':x,'n':k,'by_sym':{...}, ['usd','inr']}}"""
    def fy_sort(label):
        return int(label[2:4])
    fy_list = []
    all_sym = {}
    for fy in sorted(per_fy, key=fy_sort):
        d = per_fy[fy]
        win, los = top_movers(d["by_sym"], 3, usd)
        entry = {"label": fy, "amt": round(d["amt"], 2) if usd else round(d["amt"]),
                 "n": d["n"], "winners": win, "losers": los}
        fy_list.append(entry)
        for s, a in d["by_sym"].items():
            all_sym[s] = all_sym.get(s, 0.0) + a
    gwin, glos = top_movers(all_sym, 3, usd)
    latest = fy_list[-1] if fy_list else None
    block = {
        "asOf": fmt_asof(asof),
        "source": source,
        "total": round(sum(e["amt"] for e in fy_list), 2) if usd else round(sum(e["amt"] for e in fy_list)),
        "ytdLabel": latest["label"] if latest else None,
        "fy": fy_list, "winners": gwin, "losers": glos,
    }
    if usd:
        block["ytdUsd"] = latest["amt"] if latest else 0
    else:
        block["ytd"] = latest["amt"] if latest else 0
    return block

def main():
    accounts, fno_corro = [], {}
    dhan = vested = groww = None
    upstox_by_fy = {}                 # dedupe the many duplicate per-FY zips
    files = sorted(glob.glob(os.path.join(REPORTS, "*")))
    for p in files:
        base = os.path.basename(p)
        ext = base.lower().rsplit(".", 1)[-1]
        try:
            if base.startswith("taxpnl-") and ext == "xlsx":
                r = parse_zerodha(p)
                if r:
                    accounts.append(r)
            elif base.startswith("FYERS_tax_pnl") and ext == "csv":
                accounts.append(parse_fyers(p))
            elif base == "TAX_PNL_REPORT.xls":
                dhan = parse_dhan(p)
            elif base.startswith("Profit-Loss Statement") and ext == "xlsx":
                vested = parse_vested(p)
            elif base.startswith("realizedPnL_") and ext == "zip":   # Upstox (7BB93B), F&O
                r = parse_upstox_zip(p)
                if r and r["fy"] and r["fy"] not in upstox_by_fy:
                    upstox_by_fy[r["fy"]] = r
            elif base.startswith("AG4907") and ext == "csv":          # Astha, old F&O
                accounts.append(parse_astha(p))
            elif base.startswith("Stocks_PnL_Report") and ext == "xlsx":  # Groww, self equity
                groww = parse_groww(p)
        except Exception as e:
            print(f"  ! {base}: {type(e).__name__}: {e}")
    accounts.extend(upstox_by_fy.values())

    # global as-of = latest exit seen anywhere
    asof = None
    for a in accounts:
        le = a.get("lastExit")
        if le and (asof is None or le > asof):
            asof = le
    if vested and vested["lastExit"] and (asof is None or vested["lastExit"] > asof):
        asof = vested["lastExit"]

    # ── INDIAN_REALIZED: mom's Zerodha (GWS919) + the user's own equity ───────
    # Folds in YXA918 (Zerodha self, equity sleeve only — its F&O stays in trading)
    # and Groww (self equity). Upstox swing = held (₹0). Symbols aggregate across all.
    indian_per_fy = {}
    def add_indian(fy, amt, n, by_sym):
        d = indian_per_fy.setdefault(fy, {"amt": 0.0, "n": 0, "by_sym": {}})
        d["amt"] += amt
        d["n"] += n
        for sym, a in by_sym.items():
            d["by_sym"][sym] = d["by_sym"].get(sym, 0.0) + a
    for a in accounts:
        if (a["sleeve"] == "indian_equity" or a["label"] == "zerodha_self") and a.get("equity"):
            add_indian(a["fy"], a["equity"]["realized"], a.get("n", 0), a.get("by_sym", {}))
    if groww:
        for fy, d in groww["by_fy"].items():
            add_indian(fy, round(d["amt"]), len(d["exits"]), d["by_sym"])
    indian_realized = realized_block(
        indian_per_fy, asof, "Zerodha (mom + self) + Groww · equity realized") if indian_per_fy else None

    # ── US_REALIZED: Vested ──────────────────────────────────────────────────
    us_realized = None
    if vested:
        us_per_fy = {}
        for fy, b in vested["by_fy"].items():
            us_per_fy[fy] = {"amt": b["usd"], "n": len(b["exits"]), "by_sym": b["sym"]}
        us_realized = realized_block(us_per_fy, vested["lastExit"],
                                     "Vested realized P&L · lot-level", usd=True)
        inr_by_fy = {fy: round(b["inr"]) for fy, b in vested["by_fy"].items()}
        for e in us_realized["fy"]:
            e["amtInr"] = inr_by_fy.get(e["label"])

    # ── F&O realized: per-FY × broker (GROSS, like the equity panel) + all-time ─
    fno_by_fy = {}     # fy -> {broker: gross}
    for a in accounts:
        if a["sleeve"] == "trading" and a.get("fno") and a.get("fy"):
            g = a["fno"].get("gross", a["fno"]["realized"])
            if g is not None:
                fno_by_fy.setdefault(a["fy"], {})[a["label"]] = \
                    fno_by_fy.get(a["fy"], {}).get(a["label"], 0) + g
    if dhan:
        for fy, v in dhan["fnoByFY"].items():
            fno_by_fy.setdefault(fy, {})["dhan"] = v["gross"]
    brokers_seen = sorted({b for v in fno_by_fy.values() for b in v})
    fno_realized = {
        "asOf": fmt_asof(asof),
        "source": "Broker F&O realized · gross (FIFO), by sell-date FY",
        "total": round(sum(sum(v.values()) for v in fno_by_fy.values())),
        "brokers": brokers_seen,
        "byBroker": {b: round(sum(fno_by_fy[fy].get(b, 0) for fy in fno_by_fy)) for b in brokers_seen},
        "fy": [{"label": fy, "amt": round(sum(fno_by_fy[fy].values())),
                "byBroker": {b: round(fno_by_fy[fy][b]) for b in brokers_seen if b in fno_by_fy[fy]}}
               for fy in sorted(fno_by_fy, key=lambda x: int(x[2:4]))],
    } if fno_by_fy else None

    # ── F&O DAILY: per-trade exits bucketed by exact sell-date, for the Trading-tab
    # calendar backfill. Only brokers whose report carries a per-trade table land
    # here (Dhan today); FY-only brokers stay in fno_realized above. sleeve = S01
    # for Dhan (matches the live sync's mapping). scripts/backfill-fno-ledger.mjs
    # upserts these into data/fno-ledger.json. ──
    # Dhan = S01; Fyers + the user's own Zerodha (Kite) F&O = S02. (Astha + Upstox
    # exports carry no per-trade trade-date, so they can't be bucketed daily and
    # stay in fno_realized above.) Per-FY brokers (fyers, zerodha_self) contribute
    # a flat date→P&L dict; Dhan gives gross+net per day. Aggregate across all FYs.
    DAILY_SLEEVE = {"dhan": "S01", "fyers": "S02", "zerodha_self": "S02", "upstox": "S02"}
    daily = {}                                  # (date, broker) -> {gross, net}
    if dhan and dhan.get("fnoByDay"):
        for d, v in dhan["fnoByDay"].items():
            daily[(d, "dhan")] = {"gross": v["gross"], "net": v["net"]}
    for a in accounts:                          # fyers + zerodha_self (sleeve 'trading')
        if a.get("sleeve") == "trading" and a.get("fnoByDay") and a["label"] in DAILY_SLEEVE:
            for d, amt in a["fnoByDay"].items():
                cur = daily.setdefault((d, a["label"]), {"gross": 0.0, "net": 0.0})
                cur["gross"] += amt
                cur["net"] += amt              # no per-trade charge split → net == gross
    fno_daily = [{"date": d, "broker": b, "sleeve": DAILY_SLEEVE[b],
                  "gross": round(v["gross"], 2), "net": round(v["net"], 2)}
                 for (d, b), v in daily.items()]
    fno_daily.sort(key=lambda x: (x["date"], x["broker"]))

    out = {
        "generatedFrom": "data/reports/* (broker tax/P&L exports)",
        "asOf": fmt_asof(asof),
        "note": "PII-stripped, derived — regen with scripts/parse-broker-tax.py; never hand-edit.",
        "accounts": [
            {"label": a["label"], "broker": a["label"].split("_")[0], "owner": a["owner"],
             "sleeve": a["sleeve"], "fy": a["fy"],
             "equity": a.get("equity"), "fno": a.get("fno"), "n": a.get("n")}
            for a in accounts
        ],
        "indian_realized": indian_realized,
        "us_realized": us_realized,
        "fno_realized": fno_realized,
        "fno_daily": fno_daily,
        "dhan_allTime": dhan["allTime"] if dhan else None,
        "groww_alltime": groww["allTime"] if groww else None,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f"wrote {OUT}  (asOf {out['asOf']})")
    return out

if __name__ == "__main__":
    main()
