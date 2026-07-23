#!/usr/bin/env python3
"""
parse-vested.py — convert the Vested / DriveWealth transactions export
(Vested_Transactions*.xlsx) into the two US-sleeve stores it feeds:

  * data/us_trades.json  — per-symbol USD-flow + cash replay the US historical
                           growth curve is reconstructed from (app/lib/backfill.js)
  * US_DIVIDENDS block in data/portfolio.private.json — the "Dividend Income" card
                           on the US tab (app/components/tabs/USTab.js); a
                           FULL-REPLACE block regenerated each run (like BASIC_PAY)

Mirrors scripts/parse-payslip.py's CLI:
  (default)              review mode — parse + print a PII-free summary, NO writes
  --one <file> [--porc.] single-file probe for the ingest registry wrapper
                         (scripts/ingest/parsers/vested.mjs); prints JSON status
                         + naturalKey (the export's latest activity date)
  --write                rebuild BOTH stores from the UNION of every export in the
                         append-corpus (data/reports/vested/*.xlsx)

CUMULATIVE (append-corpus): the export is uploaded MONTH-BY-MONTH, not as one cumulative
file, so we KEEP every upload and rebuild from the union of all rows — deduped by full-row
fingerprint (an identical row in two overlapping exports counts ONCE; a genuinely repeated
row within one export is preserved). A month-only upload therefore accumulates instead of
wiping prior history. Cash EOD per date = the newest export's balance for that date.

ENV OVERRIDES (for testing to copies, live untouched):
  VESTED_DIR      append-corpus dir         (default data/reports/vested/) — all *.xlsx unioned
  VESTED_XLSX     legacy single export      (default data/reports/Vested_Transactions.xlsx) —
                  ALWAYS unioned as the permanent historical baseline when present (so the
                  pre-migration full history survives the first month-only upload)
  US_TRADES_OUT   us_trades.json out path   (default data/us_trades.json)
  VESTED_PRIVATE  portfolio.private.json    (default data/portfolio.private.json) —
                  read for the flows/other split AND patched with US_DIVIDENDS

us_trades.json SHAPE (byte-compatible with the existing file):
  {"asOf": "YYYY-MM-DD",
   "cash":  {date: usd_balance, ...},              # broker cash balance, end of day
   "flows": {SYMBOL: [[date, usd], ...], ...},     # net USD into/out of each HELD
   "other": [[date, usd], ...]}                    # same, over every NON-held symbol
  flows vs other split: a symbol lands in `flows` iff it is a CURRENT US holding
  (portfolio.private.json -> US[].sym); everything else folds into `other`. Buy =
  +Cash Amount, Sell = -Cash Amount; same-day same-symbol summed; 0.00 dropped.
  cash = the "Account Balance" column at each date's latest transaction (true EOD,
  already reflecting deposits, trades, dividends, taxes and fees).

US_DIVIDENDS block (Method A — dividends only):
  from the Income sheet's Dividend rows ONLY (Tax rows -> taxAllTime; the
  ticker-less "Balance" adjustment rows are EXCLUDED — they are not dividends).
  {asOf, grossAllTime, taxAllTime, netAllTime, last12Gross, top[6], fy[]} where
  the per-symbol `top` and per-FY `fy` breakdowns both sum to grossAllTime.

PII-SAFE: reads only Trades / All Transactions / Transfers / Income and emits ONLY
tickers, USD amounts and USD balances. The "My Account" sheet (name / email /
account number) and the security-name comments are never read into any output.
"""
import os, re, sys, json
import datetime as _dt
from collections import defaultdict

REPORTS = os.path.join("data", "reports")
DEFAULT_IN = os.environ.get("VESTED_XLSX", os.path.join(REPORTS, "Vested_Transactions.xlsx"))
DEFAULT_OUT = os.environ.get("US_TRADES_OUT", os.path.join("data", "us_trades.json"))
PRIVATE = os.environ.get("VESTED_PRIVATE", os.path.join("data", "portfolio.private.json"))
# Vested HOLDINGS export (positions snapshot) — the authoritative source for the
# US[] composition (qty/cost/inv). A SEPARATE file from the transactions export;
# see the --holdings mode. Reconstruction from the transactions Trades sheet is
# NOT viable (splits mint shares with no buy row; DRIP shares aren't booked as
# trades) — this positions file gives current qty + avg cost + invested directly.
HOLDINGS_IN = os.environ.get("VESTED_HOLDINGS_XLSX", os.path.join(REPORTS, "Vested_Holdings.xlsx"))

# First-appearance curation for holdings the current US[] has never seen. name/cat
# are USER-CURATED (cat drives CAT_COLORS + the allocation mix; it is not in any
# Vested export) so a genuinely new ticker needs a human category ONCE. After the
# first successful write the symbol lives in US[] and its name/cat are preserved
# from there, so an entry here can be pruned. A new ticker absent from BOTH US[]
# and this map FAILS the write (refuse to ship an uncategorised holding).
NEW_META = {
    "BMNR": {"name": "Bitmine Immersion", "cat": "Crypto"},   # crypto miner (mirrors HUT/IREN/CIFR…)
}

_TIME_RE = re.compile(r"^\s*(\d{1,2}):(\d{2}):(\d{2})\s*([AP]M)\s*$", re.I)


def num(x):
    try:
        return float(str(x).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def day(x):
    """First 10 chars of a cell -> 'YYYY-MM-DD' (dates come through as datetime)."""
    return str(x)[:10] if x is not None else None


def _secs(tm):
    """'07:06:19 PM' -> seconds since midnight; None if unparseable."""
    m = _TIME_RE.match(str(tm or ""))
    if not m:
        return None
    h, mi, s, ap = int(m.group(1)), int(m.group(2)), int(m.group(3)), m.group(4).upper()
    if ap == "PM" and h != 12:
        h += 12
    if ap == "AM" and h == 12:
        h = 0
    return h * 3600 + mi * 60 + s


def _fy_start(dstr):
    """Indian FY (Apr-Mar) start year for a 'YYYY-MM-DD' date."""
    y, m = int(dstr[:4]), int(dstr[5:7])
    return y if m >= 4 else y - 1


def _minus_year(dstr):
    """dstr minus one calendar year (Feb-29 -> Feb-28), as ISO."""
    y, m, d = int(dstr[:4]), int(dstr[5:7]), int(dstr[8:10])
    try:
        return _dt.date(y - 1, m, d).isoformat()
    except ValueError:
        return _dt.date(y - 1, m, d - 1).isoformat()


def load_held():
    """Current US holdings tickers — the flows/other split key. None if unavailable."""
    if not os.path.exists(PRIVATE):
        return None
    try:
        priv = json.load(open(PRIVATE, encoding="utf-8"))
    except (ValueError, OSError):
        return None
    us = priv.get("US")
    if not isinstance(us, list):
        return None
    return {h["sym"] for h in us if isinstance(h, dict) and h.get("sym")}


def build_dividends(div, tax_sum, asof, topn=6):
    """Method A US_DIVIDENDS block from Income Dividend rows (`div` = [(date, sym,
    usd)]) + summed tax. Empty {} if the export has no dated activity."""
    if not asof:
        return {}
    gross = round(sum(a for _, _, a in div), 2)
    taxall = round(abs(tax_sum), 2)
    net = round(gross - taxall, 2)
    cut = _minus_year(asof)
    last12 = round(sum(a for d, _, a in div if d > cut), 2)

    per = defaultdict(float)
    for _d, tk, a in div:
        per[tk] += a
    top = [{"sym": s, "amt": round(v, 2)}
           for s, v in sorted(per.items(), key=lambda kv: (-kv[1], kv[0]))[:topn]]

    fyagg = defaultdict(float)
    for d, _tk, a in div:
        fyagg[_fy_start(d)] += a
    fy = [{"label": f"FY{str(s)[2:]}-{str(s + 1)[2:]}", "amt": round(fyagg[s], 2)}
          for s in sorted(fyagg)]

    return {
        "asOf": _dt.date.fromisoformat(asof).strftime("%d %b %Y"),
        "grossAllTime": gross, "taxAllTime": taxall, "netAllTime": net,
        "last12Gross": last12, "top": top, "fy": fy,
    }


# ── cumulative append-corpus (union of ALL exports, deduped) ──────────────────
# The Vested export is uploaded MONTH-BY-MONTH, not as one cumulative file, so
# rebuilding from a single xlsx would wipe prior history. Instead we KEEP every
# export in the corpus dir (VESTED_DIR) and rebuild us_trades.json from the UNION
# of all rows — deduped by a full-row fingerprint so an identical row in two
# overlapping exports counts ONCE (same txn), while a genuinely repeated row within
# a single export is preserved (multiplicity = the max count seen in any one file).
_NCOLS = {"Trades": 10, "Income": 5, "Transfers": 4}   # row widths read below (for the fingerprint)


def _norm_cell(c):
    if c is None:
        return ""
    if isinstance(c, bool):
        return str(c)
    if isinstance(c, (int, float)):
        return f"{float(c):.6f}"                        # stable numeric key (avoids repr drift)
    return str(c).strip()


def _fp(row, ncols):
    """Full-row fingerprint: two byte-identical rows across overlapping exports collapse
    to one txn; any differing field (time/qty/price/amount) keeps them distinct."""
    return tuple(_norm_cell(c) for c in row[:ncols])


def _data_rows(wb, sheet):
    """(index, row) for data rows — skip the header (row 0) and fully-blank rows."""
    if sheet not in wb.sheetnames:
        return
    for i, r in enumerate(wb[sheet].iter_rows(values_only=True)):
        if i == 0 or not any(c is not None for c in r):
            continue
        yield i, r


def collect(paths):
    """Union the transaction rows across ALL corpus exports (append-corpus, like the
    payslips). Row-union sheets (Trades/Income/Transfers) dedupe by fingerprint keeping
    per-file multiplicity (max across files). Cash EOD is computed PER FILE with the
    original latest-time / top-most-on-tie rule, then merged newest-file-wins per date
    (a later upload is the freshest balance for a shared date). Empty corpus -> empty."""
    import openpyxl  # lazy — only --one/--write need it, keeps import errors local
    union = {sh: {} for sh in _NCOLS}                  # sh -> {fp: (count, rep_row)}
    cash = {}                                          # date -> (file_rank, within_key, balance)
    maxd = ""
    for rank, p in enumerate(sorted(paths)):           # sorted => later filename (later asOf) is newer
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        for sh, nc in _NCOLS.items():
            local = {}                                 # per-file counts (so overlaps don't inflate)
            for _i, r in _data_rows(wb, sh):
                d = day(r[0])
                if d:
                    maxd = max(maxd, d)
                k = _fp(r, nc)
                local[k] = (local.get(k, (0, r))[0] + 1, r)
            cur = union[sh]
            for k, (cnt, r) in local.items():
                if k not in cur or cnt > cur[k][0]:    # keep the MAX multiplicity across files
                    cur[k] = (cnt, r)
        best = {}                                      # this file's EOD cash (original tie-break)
        for i, r in _data_rows(wb, "All Transactions"):
            date, tm, _typ, _amt, bal, _comment = r[:6]
            if bal is None:
                continue
            d = day(date)
            maxd = max(maxd, d)
            key = (_secs(tm) if _secs(tm) is not None else -1, -i)
            if d not in best or key > best[d][0]:
                best[d] = (key, num(bal))
        for d, (wk, bal) in best.items():
            if d not in cash or rank >= cash[d][0]:    # newer file wins a shared date
                cash[d] = (rank, wk, bal)
        wb.close()
    rows = {sh: [] for sh in _NCOLS}
    for sh in _NCOLS:
        for _k, (cnt, r) in union[sh].items():
            rows[sh].extend([r] * cnt)
    return {"trades": rows["Trades"], "income": rows["Income"], "transfers": rows["Transfers"],
            "cash": {d: round(v[2], 2) for d, v in sorted(cash.items())}, "asOf": maxd}


def corpus_paths():
    """The full append-corpus: every export in VESTED_DIR PLUS the legacy single
    VESTED_XLSX baseline when present. The legacy file is ALWAYS included (not just as
    an empty-dir fallback) so the pre-migration full history stays in the union once the
    corpus dir gets its first month-only upload — otherwise that upload would drop it.
    Dedup collapses any overlap between the baseline and a re-uploaded month."""
    d = os.environ.get("VESTED_DIR", os.path.join(REPORTS, "vested"))
    ps = []
    if os.path.isdir(d):
        ps = sorted(os.path.join(d, f) for f in os.listdir(d) if f.lower().endswith(".xlsx"))
    if os.path.exists(DEFAULT_IN) and DEFAULT_IN not in ps:
        ps.append(DEFAULT_IN)                          # permanent historical baseline
    return ps


def _reduce(c, held):
    """Aggregate the unioned corpus rows -> {asOf, cash, flows, other, cashflows,
    dividends}. Same reduction as the original single-file parse, over the union."""
    held = held or set()
    flows = defaultdict(lambda: defaultdict(float))    # sym -> date -> usd
    other = defaultdict(float)                          # date -> usd

    # Trades -> per-symbol / other net USD (Buy +, Sell -)
    for r in c["trades"]:
        date, _tm, _name, tk, act, _ot, _qty, _pps, cash, _comm = r[:10]
        if tk is None or cash is None:
            continue
        v = num(cash) * (1 if str(act).lower().startswith("buy") else -1)
        (flows[tk] if tk in held else other)[day(date)] += v

    # Income -> dividends (per-symbol) + withholding tax (Method A; ticker-less
    # "Balance" adjustment rows are NOT dividends and are excluded).
    div = []
    tax_sum = 0.0
    for r in c["income"]:
        date, _tm, act, tk, amt = r[:5]
        if amt is None:
            continue
        if "tax" in str(act).lower():
            tax_sum += num(amt)
        elif tk is not None:
            div.append((day(date), tk, num(amt)))

    # Transfers -> US_CASHFLOWS deposit/withdrawal ledger (Deposit +, Withdrawal -).
    cf = defaultdict(float)
    for r in c["transfers"]:
        date, _tm, act, amt = r[:4]
        if amt is None:
            continue
        cf[day(date)] += num(amt) * (1 if str(act).lower().startswith("dep") else -1)
    cashflows = [{"date": d, "invested": round(cf[d], 2)}
                 for d in sorted(cf) if round(cf[d], 2) != 0.0]

    # emit us_trades — symbols alphabetical, dates ascending, drop 0.00 entries
    out_flows = {}
    for s in sorted(flows):
        pts = [[d, round(flows[s][d], 2)] for d in sorted(flows[s]) if round(flows[s][d], 2) != 0.0]
        if pts:
            out_flows[s] = pts
    out_other = [[d, round(other[d], 2)] for d in sorted(other) if round(other[d], 2) != 0.0]

    return {"asOf": c["asOf"], "cash": c["cash"], "flows": out_flows, "other": out_other,
            "cashflows": cashflows, "dividends": build_dividends(div, tax_sum, c["asOf"])}


def parse(path, held):
    """Back-compat single-file parse (one export) — used by review of a lone file / tests."""
    return _reduce(collect([path]), held)


def write_json(obj, path, compact):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        if compact:
            json.dump(obj, f, separators=(",", ":"), ensure_ascii=False)
        else:
            json.dump(obj, f, indent=1, ensure_ascii=False)


# ── Holdings (positions) export → US[] composition ────────────────────────────

def _hdr_index(row):
    """Map a Holdings header row to column indices by fuzzy name (order-independent).
    First match wins per field; if/elif keeps 'Ticker' from also claiming 'name'."""
    idx = {}
    for i, c in enumerate(row):
        k = str(c or "").lower()
        if "ticker" in k and "sym" not in idx: idx["sym"] = i
        elif "name" in k and "name" not in idx: idx["name"] = i
        elif "shares" in k and "qty" not in idx: idx["qty"] = i
        elif "average cost" in k and "cost" not in idx: idx["cost"] = i
        elif "amount invested" in k and "inv" not in idx: idx["inv"] = i
        elif "current value" in k and "val" not in idx: idx["val"] = i
    return idx


def parse_holdings(path):
    """Reduce a Vested Holdings export to [{sym, name(file), qty, cost, inv, val}]
    from the `Holdings` sheet, plus the as-of date from `User Details`. PII-SAFE:
    reads ONLY the Holdings sheet + the 'Period' cell of User Details (never the
    User/Govt Id/Account/Email columns)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Holdings" not in wb.sheetnames:
        raise ValueError("no 'Holdings' sheet — not a Vested Holdings export")

    asof = None
    if "User Details" in wb.sheetnames:
        rows = list(wb["User Details"].iter_rows(values_only=True))
        if len(rows) > 1 and rows[1] and rows[1][0]:            # 'Period' column only
            m = re.search(r"(\d{2} \w{3} \d{4})", str(rows[1][0]))   # "As of 06 Jul 2026"
            if m:
                try:
                    asof = _dt.datetime.strptime(m.group(1), "%d %b %Y").date().isoformat()
                except ValueError:
                    asof = m.group(1)

    rows = list(wb["Holdings"].iter_rows(values_only=True))
    idx = _hdr_index(rows[0]) if rows else {}
    for req in ("sym", "qty", "cost", "inv"):
        if req not in idx:
            raise ValueError(f"Holdings sheet missing a '{req}' column")
    out = []
    for r in rows[1:]:
        sym = r[idx["sym"]]
        if not sym:
            continue
        out.append({
            "sym": str(sym).strip(),
            "name": (str(r[idx["name"]]).strip() if "name" in idx and r[idx["name"]] else None),
            "qty": num(r[idx["qty"]]),
            "cost": round(num(r[idx["cost"]]), 2),
            "inv": round(num(r[idx["inv"]]), 2),
            "val": round(num(r[idx["val"]]), 2) if "val" in idx else None,
        })
    return {"asOf": asof, "holdings": out}


def build_us(holdings, cur_us):
    """Merge a holdings snapshot into US[] rows. Numeric fields (qty/cost/inv) come
    from the broker snapshot; curated name/cat are PRESERVED from the current US[]
    (or NEW_META for a first-seen ticker). Existing order is kept; new tickers are
    appended; tickers no longer held are dropped. Returns (rows, new_syms, dropped,
    uncategorised)."""
    hold = {h["sym"]: h for h in holdings}
    curmap = {h["sym"]: h for h in cur_us}

    uncategorised = []
    def meta(sym, file_name):
        if sym in curmap:
            return curmap[sym].get("name"), curmap[sym].get("cat")
        if sym in NEW_META:
            return NEW_META[sym]["name"], NEW_META[sym]["cat"]
        uncategorised.append(sym)
        return (file_name or sym), None

    def row(sym):
        h = hold[sym]
        name, cat = meta(sym, h.get("name"))
        return {"sym": sym, "name": name, "cat": cat,
                "qty": h["qty"], "cost": h["cost"], "inv": h["inv"]}

    rows = [row(h["sym"]) for h in cur_us if h["sym"] in hold]     # keep order
    new_syms = [s for s in hold if s not in curmap]
    rows += [row(s) for s in new_syms]                            # append newcomers
    dropped = [s for s in curmap if s not in hold]
    return rows, new_syms, dropped, uncategorised


def probe(path):
    """Registry probe: latest activity date + coarse counts. No PII, no split. One
    workbook open; maxd scans the same sheets parse() does so keys agree."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    maxd = ""
    tickers = set()
    trades = 0
    cash_days = set()
    if "Trades" in wb.sheetnames:
        for i, r in enumerate(wb["Trades"].iter_rows(values_only=True)):
            if i == 0:
                continue
            d = day(r[0])
            if d:
                maxd = max(maxd, d)
            if r[3] is not None:
                tickers.add(r[3]); trades += 1
    for sh in ("All Transactions", "Income", "Transfers"):
        if sh in wb.sheetnames:
            for i, r in enumerate(wb[sh].iter_rows(values_only=True)):
                if i == 0:
                    continue
                d = day(r[0])
                if d:
                    maxd = max(maxd, d)
                    if sh == "All Transactions" and r[4] is not None:
                        cash_days.add(d)
    return {"asOf": maxd or None, "key": maxd or None,
            "tickers": len(tickers), "trades": trades, "cashDays": len(cash_days)}


def _run_holdings():
    """--holdings modes: probe (--one), review (default), or write US[] (--write)."""
    # probe — single-file porcelain for scripts/ingest/parsers/vested-holdings.mjs
    if "--one" in sys.argv:
        path = sys.argv[sys.argv.index("--one") + 1]
        try:
            hp = parse_holdings(path)
        except Exception as e:  # noqa: BLE001 — a foreign/broken xlsx is a FAIL, not a crash
            print(json.dumps({"asOf": None, "error": f"{type(e).__name__}: {e}"[:200]}))
            sys.exit(1)
        st = {"asOf": hp["asOf"], "key": hp["asOf"], "holdings": len(hp["holdings"])}
        print(json.dumps(st))
        sys.exit(0 if hp["holdings"] else 1)

    hp = parse_holdings(HOLDINGS_IN)
    priv = json.load(open(PRIVATE, encoding="utf-8"))
    cur_us = priv.get("US", [])
    rows, new_syms, dropped, uncat = build_us(hp["holdings"], cur_us)

    # refuse to ship an uncategorised holding — a new ticker needs a name/cat in
    # US[] or NEW_META first (cat drives CAT_COLORS + the allocation mix).
    if uncat:
        print(f"FAIL: {len(uncat)} new ticker(s) with no curated cat: {', '.join(uncat)} "
              f"— add to NEW_META (or US[]) then re-run. NOT writing.")
        sys.exit(1)

    inv = round(sum(r["inv"] for r in rows), 2)
    curinv = round(sum(h.get("inv", 0) for h in cur_us), 2)
    catof = {r["sym"]: r["cat"] for r in rows}
    print(f"holdings: {HOLDINGS_IN}  asOf {hp['asOf']}")
    print(f"US[]: {len(cur_us)} -> {len(rows)} holdings | invested ${curinv} -> ${inv}")
    if new_syms:
        print("           NEW: " + ", ".join(f"{s} ({catof[s]})" for s in new_syms))
    if dropped:
        print("           SOLD OUT (dropped): " + ", ".join(dropped))

    if "--write" not in sys.argv:
        print("(review only — pass --write to update US[] + re-seed KV)")
        return

    priv["US"] = rows
    write_json(priv, PRIVATE, compact=False)
    print(f"wrote US[]: {len(rows)} holdings, invested ${inv} -> {PRIVATE} (re-seed KV to publish)")


def main():
    if "--holdings" in sys.argv:
        _run_holdings()
        return

    # --one <file> [--porcelain]: single-file probe for scripts/ingest/parsers/vested.mjs
    if "--one" in sys.argv:
        path = sys.argv[sys.argv.index("--one") + 1]
        try:
            st = probe(path)
        except Exception as e:  # noqa: BLE001 — a broken/foreign xlsx is a FAIL, not a crash
            print(json.dumps({"asOf": None, "error": f"{type(e).__name__}: {e}"[:200]}))
            sys.exit(1)
        print(json.dumps(st))
        sys.exit(0 if st.get("asOf") else 1)

    held = load_held()

    if "--write" in sys.argv:
        if not held:
            # the flows/other split is essential — refuse rather than dump every
            # symbol into `other` (mirrors payslip's guarded-seed refusal semantics)
            print("FAIL: could not load US holdings from " + PRIVATE
                  + " (needed for the flows/other split) — refusing to write")
            sys.exit(1)
        paths = corpus_paths()                          # union of ALL exports in the append-corpus
        if not paths:
            print("FAIL: no Vested exports in the corpus (VESTED_DIR / VESTED_XLSX) — nothing to write")
            sys.exit(1)
        data = _reduce(collect(paths), held)
        if not data["asOf"]:                            # empty/no-activity corpus — don't clobber us_trades.json
            print(f"FAIL: {len(paths)} export(s) but no dated activity — refusing to overwrite {DEFAULT_OUT}")
            sys.exit(1)

        # 1) us_trades.json — asOf/cash/flows/other ONLY (never the dividends key)
        us = {"asOf": data["asOf"], "cash": data["cash"], "flows": data["flows"], "other": data["other"]}
        write_json(us, DEFAULT_OUT, compact=True)
        print(f"wrote us_trades.json: {len(paths)} export(s) in corpus, {len(us['flows'])} held symbols, "
              f"{len(us['other'])} other pts, {len(us['cash'])} cash days, asOf {us['asOf']} -> {DEFAULT_OUT}")

        # 2) US_DIVIDENDS + US_CASHFLOWS -> portfolio.private.json (full-replace
        #    blocks, like BASIC_PAY). US_CASHFLOWS = the Transfers ledger (deposits/
        #    withdrawals); it feeds the Capital Deployment card's US stream. One
        #    private-JSON read + one write patches both.
        dv = data["dividends"]
        cf = data["cashflows"]
        priv = json.load(open(PRIVATE, encoding="utf-8"))
        priv["US_DIVIDENDS"] = dv
        priv["US_CASHFLOWS"] = cf
        write_json(priv, PRIVATE, compact=False)
        print(f"wrote US_DIVIDENDS: gross {dv['grossAllTime']} / tax {dv['taxAllTime']} / net {dv['netAllTime']}, "
              f"asOf {dv['asOf']} -> {PRIVATE} (re-seed KV to publish)")
        net = round(sum(c["invested"] for c in cf), 2)
        print(f"wrote US_CASHFLOWS: {len(cf)} transfers, net ${net}, "
              f"{cf[0]['date'] if cf else '—'} .. {cf[-1]['date'] if cf else '—'} -> {PRIVATE} (re-seed KV to publish)")
        return

    # review mode — parse + PII-free summary, no writes
    paths = corpus_paths()
    if not paths:
        print("(no Vested exports in the corpus — upload one to seed VESTED_DIR)")
        return
    data = _reduce(collect(paths), held or set())
    dv = data["dividends"]
    print(f"corpus: {len(paths)} export(s) in {os.environ.get('VESTED_DIR', os.path.join(REPORTS, 'vested'))}")
    print(f"asOf (latest activity): {data['asOf']}")
    print(f"us_trades: {len(data['flows'])} held symbols in flows"
          + ("" if held else "  [WARN: no holdings list -> everything in 'other']")
          + f", {len(data['other'])} other pts, {len(data['cash'])} cash days")
    if data["cash"]:
        d = sorted(data["cash"])
        print(f"           cash {d[0]} .. {d[-1]}, last balance ${data['cash'][d[-1]]}")
    cf = data["cashflows"]
    net = round(sum(c["invested"] for c in cf), 2)
    print(f"US_CASHFLOWS: {len(cf)} transfers (deposits/withdrawals), net ${net}"
          + (f", {cf[0]['date']} .. {cf[-1]['date']}" if cf else ""))
    if held:
        missing = sorted(held - set(data["flows"]))
        print(f"           holdings covered: {len(set(data['flows']) & held)}/{len(held)}"
              + (f"  missing: {missing}" if missing else ""))
    if dv:
        fysum = round(sum(f["amt"] for f in dv["fy"]), 2)
        print(f"US_DIVIDENDS: gross {dv['grossAllTime']} / tax {dv['taxAllTime']} / net {dv['netAllTime']} "
              f"/ last12 {dv['last12Gross']}, asOf {dv['asOf']}")
        print(f"           fy sums to {fysum} (ties gross: {fysum == dv['grossAllTime']}); "
              f"this FY {dv['fy'][-1]['label']} {dv['fy'][-1]['amt']}; top {dv['top'][0]['sym']} {dv['top'][0]['amt']}")


if __name__ == "__main__":
    main()
