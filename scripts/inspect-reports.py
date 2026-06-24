#!/usr/bin/env python3
"""
inspect-reports.py — print the STRUCTURE of the broker reports in data/reports/
so the daily-F&O backfill can be extended to Zerodha/Kite, Fyers, and Upstox
without guessing column positions.

PII-SAFE: it masks values — numbers become '#', account-id / PAN-looking tokens
become '<id>'. It prints only sheet names, the header/label rows, and section
titles (column names + section labels are what's needed to map columns). Dates
are kept (not personal) so the sell-date column is identifiable.

Run on the laptop (where data/reports/ lives), paste the output back:
    python scripts\\inspect-reports.py
"""
import os, re, csv, glob
import openpyxl

REPORTS = os.path.join("data", "reports")
MAXROWS = 18           # header/section rows live near the top of each section
MAXSHEETROWS = 60      # scan this many rows per sheet to surface section labels

PAN = re.compile(r"^[A-Z]{5}\d{4}[A-Z]$")
ID  = re.compile(r"^[A-Z0-9]{6,}$")          # client/UCC ids like YS59535, 7BB93B
NUMISH = re.compile(r"^-?[\d,]+\.?\d*$")
DATEISH = re.compile(r"\d{1,4}[-/]\d{1,2}[-/]\d{1,4}")

def mask(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s == "":
        return ""
    if DATEISH.search(s):
        return s                              # dates are safe + identify the date column
    if NUMISH.match(s.replace(" ", "")):
        return "#"                            # amounts/qty → masked
    if PAN.match(s) or (ID.match(s) and any(c.isdigit() for c in s)):
        return "<id>"
    return s                                  # column headers, section titles, symbols

def show_row(cells):
    out = [f"{i}:{mask(c)}" for i, c in enumerate(cells) if mask(c) != ""]
    return "  ".join(out)

def inspect_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    print(f"  sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  --- sheet '{sn}' ---")
        shown = 0
        for ri, row in enumerate(ws.iter_rows(max_row=MAXSHEETROWS)):
            vals = [c.value for c in row]
            line = show_row(vals)
            # print header-ish rows: those with >=2 string (non-#, non-date) cells
            strs = sum(1 for c in vals if mask(c) not in ("", "#") and not DATEISH.search(str(c or "")))
            if line and strs >= 2 and shown < MAXROWS:
                print(f"    r{ri}: {line}")
                shown += 1
    wb.close()

def inspect_csv(path):
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        rows = list(csv.reader(f))
    print(f"  rows: {len(rows)}")
    shown = 0
    for ri, r in enumerate(rows[:MAXSHEETROWS]):
        line = show_row(r)
        strs = sum(1 for c in r if mask(c) not in ("", "#") and not DATEISH.search(str(c or "")))
        if line and strs >= 2 and shown < MAXROWS:
            print(f"    r{ri}: {line}")
            shown += 1

def main():
    files = sorted(glob.glob(os.path.join(REPORTS, "*")))
    if not files:
        print(f"no files in {REPORTS}/ — drop your broker reports there first")
        return
    for p in files:
        print(f"\n=== {os.path.basename(p)} ===")
        try:
            ext = os.path.splitext(p)[1].lower()
            if ext in (".xlsx", ".xls", ".xlsm"):
                inspect_xlsx(p)
            elif ext == ".csv":
                inspect_csv(p)
            else:
                print(f"  (skipped — {ext})")
        except Exception as e:
            print(f"  ERROR reading: {e}")

if __name__ == "__main__":
    main()
